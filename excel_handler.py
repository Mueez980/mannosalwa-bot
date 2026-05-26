import openpyxl
import os
from datetime import datetime

EXCEL_FILE = "orders.xlsx"

def setup_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders"
        headers = [
            "Order Number",
            "Date and Time",
            "Customer Name",
            "Phone Number",
            "Delivery Address",
            "Items Ordered",
            "Special Instructions",
            "Status"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(
                fill_type="solid", fgColor="1B5E20")
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 15
        wb.save(EXCEL_FILE)

def save_order(name, phone, address, items, special="None"):
    setup_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    order_num = ws.max_row
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.append([
        order_num,
        now,
        name,
        phone,
        address,
        items,
        special,
        "Pending"
    ])
    for col in range(1, 9):
        ws.cell(row=ws.max_row, column=col).fill = \
            openpyxl.styles.PatternFill(
                fill_type="solid", fgColor="F1F8E9")
    wb.save(EXCEL_FILE)
    print(f"Order {order_num} saved!")
    return order_num
